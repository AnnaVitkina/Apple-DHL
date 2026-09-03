"""
Build a matrix workbook from extracted DHL processing dataframes.

Matrix layout:
  - Shipment info (lane details)
  - Transport cost columns with multi-row headers (weight brackets)
"""

from __future__ import annotations

import re
import sys
from datetime import date, datetime
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from project_paths import INPUT_DIR, OUTPUT_DIR, PROCESSING_DIR, ensure_workspace_dirs

EXTRACTED_GLOB = "*_extracted.xlsx"
RA_SUBDIR = "RA"
EXCEL_SUFFIXES = {".xlsx", ".xls", ".xlsm"}
ISO_COLUMN_PATTERN = re.compile(r"^[A-Z]{2}$")
PUK_FILE_PATTERN = re.compile(r"DHLPUK", re.IGNORECASE)
PUK_POSTAL_ZONE_COLUMN_PATTERN = re.compile(r"^[A-Z]{2}_[A-Z0-9]+$")
EU_FR_POSTAL_FILE_PATTERN = re.compile(
    r"DHL_.*_EU_(?:AC|FG(?:_TDD)?)",
    re.IGNORECASE,
)
TIME_DEFINITE_PATTERN = re.compile(r"time\s+definite", re.IGNORECASE)
FR_TIME_DEFINITE_POSTAL_CODE = "98000"
WEIGHT_RANGE_PATTERN = re.compile(r"^(\d+)\s*(?:to|-)\s*(\d+)$", re.IGNORECASE)

SHIPMENT_COLUMNS = (
    "Lane #",
    "Tab",
    "Origin country",
    "Shipment Type",
    "Service",
    "Service level",
    "Destination country",
    "Destination Postal Code",
    "Carrier Account number",
    "Business Segment",
    "Category",
    "Shipping Condition",
    "Valid from",
    "Valid to",
)

CURRENCY_COLUMN = "Currency"
TAB_INDEX_SHEET = "Tab Index"
RETURN_DESCRIPTION_PATTERN = re.compile(r"shipments\s+back", re.IGNORECASE)
RETURN_CATEGORY = "return"

TRANSPORT_COST_GROUP = "Transport cost"
SFS_TRANSPORT_COST_GROUP = "Transport cost (SFS)"
SFS_TAB_NAME = "SFS"
EXCLUDED_RATE_TABS = {"tab index", "accessorials"}
BILLING_BS_GLOB = "billing-bs*"
RA_RATE_CARD_SHEET = "rate card"
RA_HEADER_SCAN_ROWS = 25
RA_ORIGIN_HEADER_ALIASES = frozenset({"origin country", "origin"})
RA_DESTINATION_HEADER_ALIASES = frozenset(
    {"destination country", "destination", "destination iso"}
)
RA_FILL_FIELDS = frozenset({"Category", "Shipping Condition"})
RA_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "Tab": ("tab", "tab name", "tab-name"),
    "Origin country": ("origin country", "origin"),
    "Destination country": ("destination country", "destination", "destination iso"),
    "Destination Postal Code": (
        "destination postal code",
        "postal code",
        "destination postal",
    ),
    "Shipment Type": ("shipment type",),
    "Service level": ("service level",),
    "Category": ("category",),
    "Shipping Condition": ("shipping condition",),
}
CARRIER_ACCOUNT_SPLIT_PATTERN = re.compile(r"[,/;]+")
CARRIER_ACCOUNT_TOKEN_PATTERN = re.compile(r"\b[A-Za-z0-9]+\b")
BUSINESS_SEGMENT_ABBREVIATIONS: dict[str, str] = {
    "finished goods": "FG",
    "fg": "FG",
    "applecare": "AC",
    "apple care": "AC",
    "finished goods special billing": "FGSB",
    "applecare special billing": "ACSB",
    "secureship": "SS",
    "secureship special billing": "SSSB",
    "manufacturing": "MFG",
    "seed lab": "SL",
    "corporate": "CORP",
    "apple retail": "AR",
    "charter - special billing account": "CHARTER",
    "special billing": "SB",
}

FIXED_SOURCE_COLUMNS = {
    "Version Number",
    "Version Date",
    "Destination Country",
    "Destination ISO",
    "Billing Currency",
    "Origin",
    "Shipment Type",
    "Service",
    "Service Level",
    "Chargeable Weight",
    "Rate Logic",
    "Lane Code",
}

COST_GROUP_ROW = 1
COST_NAME_ROW = 2
APPLY_IF_ROW = 3
RATE_BY_ROW = 4
COLUMN_HEADER_ROW = 5
DATA_START_ROW = 6

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
RA_FILL_HIGHLIGHT = PatternFill("solid", fgColor="FFFF00")
TRANSPORT_GROUP_FILL = PatternFill("solid", fgColor="9BC2E6")
TRANSPORT_COST_FILL = PatternFill("solid", fgColor="BDD7EE")
COST_META_FILL = PatternFill("solid", fgColor="F2F2F2")
BOLD = Font(bold=True)
NORMAL = Font()
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RATE_NUMBER_FORMAT = "#,##0.00"
THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)

EU_COUNTRIES = frozenset({
    "AT", "BE", "BG", "HR", "CY", "CZ", "DK", "EE", "FI", "FR", "DE", "GR",
    "HU", "IE", "IT", "LV", "LT", "LU", "MT", "NL", "PL", "PT", "RO", "SK",
    "SI", "ES", "SE",
})
ES_WPX_TERRITORIES = frozenset({"ES", "IC", "AD"})
DEPOT_TO_COUNTRY = {
    "LEJ": "DE",
    "CGO": "CN",
    "TYN": "CN",
    "CTU": "CN",
    "SHA": "CN",
    "PVG": "CN",
    "NKG": "CN",
    "SZX": "CN",
    "HKG": "HK",
    "7101": "IE",
}
LU_DEPOT_PATTERN = re.compile(r"^LU\d+$", re.IGNORECASE)
TAB_PREFIX_TO_COUNTRY = {
    "CN_": "CN",
    "VN_": "VN",
}
NSR_SERVICE_LEVELS = ("DOM", "ECX", "WPX")
SERVICE_LEVEL_TOKEN_PATTERN = re.compile(r"[A-Z0-9]+")
ISO_TOKEN_PATTERN = re.compile(r"\b[A-Z]{2}\b")
SERVICE_LEVEL_TO_EXP: dict[str, str] = {
    "DOM": "EXP_DOM",
    "GROUND": "EXP_DOM",
    "WPX": "EXP_WW_NONDOC",
    "ECX": "EXP_WW_EU",
    "DES": "EXP_EC_DOM",
    "ESU": "EXP_EC_EU",
    "ESI": "EXP_EC_NONDOC",
    "DOX": "EXP_WW_DOC",
    "BTC": "EXP_B2C",
    "DOT": "EXP_DOM_1200",
    "SDX": "SAMEDAY",
    "BBX": "EXP_BREAKBULK",
    "TDY": "EXP_1200_NONDOC",
    "TDT": "EXP_1200_DOC",
    "TDE": "EXP_900_NONDOC",
    "TDK": "EXP_900_DOC",
    "DOK": "EXP_DOM_900",
}
UNCONDITIONAL_EXP_SERVICES = frozenset({
    "EXP_DOM_1200",
    "EXP_1200_NONDOC",
    "EXP_1200_DOC",
    "EXP_900_NONDOC",
    "EXP_900_DOC",
    "EXP_DOM_900",
    "EXP_BREAKBULK",
    "SAMEDAY",
})


def cell_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def rate_value(value: object) -> float | int | None:
    if pd.isna(value):
        return None
    text = cell_text(value)
    if not text or text.lower() in {"on request", "n/a", "#n/a"}:
        return None
    try:
        number = float(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return None
    if number.is_integer():
        return int(number)
    return number


def format_display_date(value: object) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, datetime):
        return value.date().strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    text = cell_text(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d.%m.%Y")
        except ValueError:
            continue
    return text


def rate_logic_to_rate_by(rate_logic: object) -> str:
    text = cell_text(rate_logic).lower()
    if text == "dn rate":
        return "Flat"
    if text == "per-kg rate":
        return "p/unit"
    return ""


def _bracket_upper_label(number: float) -> str | int:
    if number == int(number):
        return int(number)
    return f"{number:g}"


def weight_bracket_label(chargeable_weight: object) -> str:
    text = cell_text(chargeable_weight)
    if not text:
        return ""

    range_match = WEIGHT_RANGE_PATTERN.match(text)
    if range_match:
        upper = _bracket_upper_label(float(range_match.group(2)))
        return f"<={upper}"

    try:
        upper = _bracket_upper_label(float(text.replace(",", "")))
    except ValueError:
        return text
    return f"<={upper}"


def bracket_sort_key(label: str) -> tuple[int, float | str]:
    match = re.match(r"^<=(.+)$", label)
    if match:
        upper_text = match.group(1)
        try:
            return (0, float(upper_text))
        except ValueError:
            return (1, upper_text)
    return (2, label)


def transport_cost_column_name(bracket_label: str) -> str:
    return f"{TRANSPORT_COST_GROUP} ({bracket_label})"


def sfs_transport_cost_column_name(bracket_label: str) -> str:
    return f"{SFS_TRANSPORT_COST_GROUP} ({bracket_label})"


def is_sfs_tab(tab_name: str) -> bool:
    return tab_name.strip().upper() == SFS_TAB_NAME


def is_standard_weight_bracket_cost_column(column_name: str) -> bool:
    prefix = f"{TRANSPORT_COST_GROUP} ("
    return column_name.startswith(prefix) and not column_name.startswith(f"{SFS_TRANSPORT_COST_GROUP} (")


def is_sfs_weight_bracket_cost_column(column_name: str) -> bool:
    return column_name.startswith(f"{SFS_TRANSPORT_COST_GROUP} (")


def is_eu_country(country_code: str) -> bool:
    return country_code.upper() in EU_COUNTRIES


def parse_service_level_tokens(service_level: object) -> list[str]:
    text = cell_text(service_level).upper()
    if not text:
        return []

    tokens: list[str] = []
    for match in SERVICE_LEVEL_TOKEN_PATTERN.findall(text):
        if match == "NSR":
            tokens.extend(NSR_SERVICE_LEVELS)
        else:
            tokens.append(match)
    return tokens


def _country_codes_in_text(value: object) -> list[str]:
    return ISO_TOKEN_PATTERN.findall(cell_text(value).upper())


def resolve_lane_country_code(
    country_value: object,
    *,
    tab_name: str = "",
    tab_index_lookup: dict[str, TabIndexInfo] | None = None,
) -> str:
    text = cell_text(country_value)
    if not text or text.lower() == "any":
        return ""

    for code in _country_codes_in_text(text):
        if ISO_COLUMN_PATTERN.match(code):
            return code

    for part in re.split(r"[,/;\s]+", text.upper()):
        if not part:
            continue
        if part in DEPOT_TO_COUNTRY:
            return DEPOT_TO_COUNTRY[part]
        if LU_DEPOT_PATTERN.match(part):
            return "LU"
        if ISO_COLUMN_PATTERN.match(part):
            return part

    tab_name = cell_text(tab_name)
    if ISO_COLUMN_PATTERN.match(tab_name):
        return tab_name

    if tab_index_lookup:
        tab_info = tab_index_lookup.get(tab_name)
        if tab_info is not None and tab_info.origins:
            resolved = resolve_lane_country_code(
                tab_info.origins,
                tab_name=tab_name,
            )
            if resolved:
                return resolved

    for prefix, country_code in TAB_PREFIX_TO_COUNTRY.items():
        if tab_name.upper().startswith(prefix):
            return country_code

    return ""


def spl_service_for_lane(origin: str, destination: str) -> str:
    if origin == "RU" or destination == "RU":
        return "CITYLINE (DHL EXP RU)"
    return "SPRINTLINE (DHL EXP NL)"


def _dom_exp_service_applicable(origin: str, destination: str) -> bool:
    return bool(origin) and origin == destination


def _ecx_exp_service_applicable(origin: str, destination: str) -> bool:
    if not origin or not destination:
        return False
    return origin != destination and is_eu_country(origin) and is_eu_country(destination)


def _wpx_exp_service_applicable(origin: str, destination: str) -> bool:
    if not destination:
        return False
    if not origin:
        return not is_eu_country(destination)
    if origin == destination:
        return False
    if not is_eu_country(origin) or not is_eu_country(destination):
        return True
    return origin in ES_WPX_TERRITORIES or destination in ES_WPX_TERRITORIES


def _esi_exp_service_applicable(origin: str, destination: str) -> bool:
    if not destination:
        return False
    if not origin:
        return not is_eu_country(destination)
    return origin != destination and (not is_eu_country(origin) or not is_eu_country(destination))


def is_exp_service_applicable(exp_service: str, origin: str, destination: str) -> bool:
    origin = origin.upper()
    destination = destination.upper()

    if exp_service == "EXP_DOM":
        return _dom_exp_service_applicable(origin, destination)
    if exp_service == "EXP_WW_EU":
        return _ecx_exp_service_applicable(origin, destination)
    if exp_service == "EXP_WW_NONDOC":
        return _wpx_exp_service_applicable(origin, destination)
    if exp_service == "EXP_EC_DOM":
        return _dom_exp_service_applicable(origin, destination)
    if exp_service == "EXP_EC_EU":
        return _ecx_exp_service_applicable(origin, destination)
    if exp_service == "EXP_EC_NONDOC":
        return _esi_exp_service_applicable(origin, destination)
    if exp_service == "EXP_WW_DOC":
        return (
            _dom_exp_service_applicable(origin, destination)
            or _ecx_exp_service_applicable(origin, destination)
            or _wpx_exp_service_applicable(origin, destination)
        )
    if exp_service == "EXP_B2C":
        return origin == "RU" and destination == "RU"
    if exp_service in UNCONDITIONAL_EXP_SERVICES:
        return True
    return False


def resolve_service_column(
    service_level: object,
    origin: str,
    destination: str,
) -> str:
    applicable: list[str] = []
    seen: set[str] = set()

    for token in parse_service_level_tokens(service_level):
        if token == "SPL":
            exp_service = spl_service_for_lane(origin, destination)
        else:
            exp_service = SERVICE_LEVEL_TO_EXP.get(token)
            if exp_service is None:
                continue

        if exp_service in seen:
            continue
        if is_exp_service_applicable(exp_service, origin, destination):
            applicable.append(exp_service)
            seen.add(exp_service)

    return "/".join(applicable)


def apply_service_column(
    matrix_df: pd.DataFrame,
    tab_index_lookup: dict[str, TabIndexInfo] | None = None,
) -> pd.DataFrame:
    if matrix_df.empty:
        return matrix_df

    result = matrix_df.copy()
    for index, row in result.iterrows():
        origin = resolve_lane_country_code(
            row.get("Origin country"),
            tab_name=cell_text(row.get("Tab")),
            tab_index_lookup=tab_index_lookup,
        )
        destination = resolve_lane_country_code(row.get("Destination country"))
        result.at[index, "Service"] = resolve_service_column(
            row.get("Service level"),
            origin,
            destination,
        )
    return result


def destination_iso_columns(df: pd.DataFrame) -> list[str]:
    return [
        str(column)
        for column in df.columns
        if str(column) not in FIXED_SOURCE_COLUMNS and ISO_COLUMN_PATTERN.match(str(column))
    ]


def is_puk_rate_file(file_path: Path | str) -> bool:
    return bool(PUK_FILE_PATTERN.search(Path(file_path).name))


def is_eu_fr_postal_duplicate_file(file_path: Path | str) -> bool:
    stem = Path(file_path).stem.replace("_extracted", "")
    return bool(EU_FR_POSTAL_FILE_PATTERN.search(stem))


def destination_rate_columns(df: pd.DataFrame, *, puk_mode: bool = False) -> list[str]:
    if puk_mode:
        return [
            str(column)
            for column in df.columns
            if str(column) not in FIXED_SOURCE_COLUMNS
        ]
    return destination_iso_columns(df)


def parse_puk_destination_column(
    column_name: str,
    *,
    default_destination_iso: str = "",
) -> tuple[str, str]:
    """Map a DHLPUK rate column to destination country and postal code zone."""
    column = cell_text(column_name).upper()
    default_destination_iso = cell_text(default_destination_iso).upper()

    if PUK_POSTAL_ZONE_COLUMN_PATTERN.match(column):
        country = column.split("_", 1)[0]
        return country, column

    if column == "IE":
        return "IE", "IE"

    if ISO_COLUMN_PATTERN.match(column):
        if column == "GB" or column == default_destination_iso:
            return column, column
        return column, ""

    if default_destination_iso:
        return default_destination_iso, column

    return column, ""


def is_excluded_matrix_tab(sheet_name: str) -> bool:
    return sheet_name.strip().lower() in EXCLUDED_RATE_TABS


def is_rate_tab(
    sheet_name: str,
    df: pd.DataFrame,
    tab_index_lookup: dict[str, TabIndexInfo] | None = None,
    *,
    puk_mode: bool = False,
) -> bool:
    if is_excluded_matrix_tab(sheet_name):
        return False
    if tab_index_lookup:
        tab_info = tab_index_lookup.get(sheet_name)
        if tab_info is None or not tab_info.is_active:
            return False
    required = {"Origin", "Chargeable Weight", "Rate Logic"}
    if not required.issubset(df.columns):
        return False
    return bool(destination_rate_columns(df, puk_mode=puk_mode))


def list_extracted_files() -> list[Path]:
    return sorted(
        [
            path
            for path in PROCESSING_DIR.glob(EXTRACTED_GLOB)
            if path.is_file() and not path.name.startswith("~$")
        ],
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )


def prompt_selection(title: str, items: list[str]) -> int:
    print(f"\n{title}")
    for index, item in enumerate(items, start=1):
        print(f"  {index}. {item}")

    while True:
        raw = input("Enter number: ").strip()
        if not raw.isdigit():
            print("Please enter a valid number.")
            continue
        choice = int(raw)
        if 1 <= choice <= len(items):
            return choice - 1
        print("Number is out of range. Try again.")


def prompt_optional_selection(title: str, items: list[str]) -> int | None:
    print(f"\n{title}")
    for index, item in enumerate(items, start=1):
        print(f"  {index}. {item}")

    while True:
        raw = input("Enter number or press Enter to skip: ").strip()
        if not raw:
            return None
        if not raw.isdigit():
            print("Please enter a valid number, or press Enter to skip.")
            continue
        choice = int(raw)
        if 1 <= choice <= len(items):
            return choice - 1
        print("Number is out of range. Try again.")


def select_extracted_file(files: list[Path], *, auto: bool = False) -> Path:
    if not files:
        print(f"No extracted files found in: {PROCESSING_DIR}")
        sys.exit(1)

    if auto or len(files) == 1:
        print(f"\nUsing extracted file: {files[0].name}")
        return files[0]

    labels = [path.name for path in files]
    return files[prompt_selection("Select extracted file to build matrix from:", labels)]


def load_rate_tabs(file_path: Path) -> list[tuple[str, pd.DataFrame]]:
    workbook = pd.ExcelFile(file_path)
    tab_index_lookup = build_tab_index_lookup(load_tab_index(file_path))
    puk_mode = is_puk_rate_file(file_path)
    loaded: list[tuple[str, pd.DataFrame]] = []

    for sheet_name in workbook.sheet_names:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        if not is_rate_tab(sheet_name, df, tab_index_lookup, puk_mode=puk_mode):
            continue
        loaded.append((sheet_name, df))

    return loaded


@dataclass(frozen=True)
class TabIndexInfo:
    tab_name: str
    is_active: bool
    has_return_lanes: bool
    billing_accounts: str
    origins: str


@dataclass(frozen=True)
class RaLaneEntry:
    service_level_tokens: frozenset[str]
    values: dict[str, str]


def normalize_billing_accounts(value: object) -> str:
    text = cell_text(value).replace("\n", " ")
    parts = [part.strip() for part in re.split(r"\s*;\s*", text) if part.strip()]
    return "; ".join(parts)


def billing_bs_file_path() -> Path | None:
    matches = sorted(INPUT_DIR.glob(BILLING_BS_GLOB))
    return matches[0] if matches else None


def ra_input_dir() -> Path:
    return INPUT_DIR / RA_SUBDIR


def list_ra_files() -> list[Path]:
    ra_dir = ra_input_dir()
    if not ra_dir.is_dir():
        return []

    return [
        path
        for path in sorted(ra_dir.iterdir())
        if path.is_file()
        and path.suffix.lower() in EXCEL_SUFFIXES
        and not path.name.startswith("~$")
    ]


def select_ra_file(files: list[Path], *, auto: bool = False) -> Path | None:
    if not files:
        return None

    if auto:
        if len(files) == 1:
            print(f"\nAuto mode: using RA file {files[0].name}")
            return files[0]
        return None

    labels = [path.name for path in files]
    choice = prompt_optional_selection(
        "Optional RA lookup file for Category / Shipping Condition:",
        labels,
    )
    if choice is None:
        print("Skipping RA lookup.")
        return None
    return files[choice]


def normalize_header_name(value: object) -> str:
    return re.sub(r"\s+", " ", cell_text(value).lower())


def find_rate_card_sheet_name(workbook: pd.ExcelFile) -> str | None:
    for sheet_name in workbook.sheet_names:
        if sheet_name.strip().lower() == RA_RATE_CARD_SHEET:
            return sheet_name
    return None


def find_ra_header_row(preview: pd.DataFrame) -> int | None:
    scan_limit = min(len(preview), RA_HEADER_SCAN_ROWS)
    for row_idx in range(scan_limit):
        labels = {
            normalize_header_name(value)
            for value in preview.iloc[row_idx]
            if not pd.isna(value)
        }
        if labels & RA_ORIGIN_HEADER_ALIASES and labels & RA_DESTINATION_HEADER_ALIASES:
            return row_idx
    return None


def read_ra_rate_card_dataframe(file_path: Path) -> tuple[pd.DataFrame | None, str]:
    try:
        workbook = pd.ExcelFile(file_path)
    except Exception as exc:
        return None, f"could not open workbook ({exc})"

    sheet_name = find_rate_card_sheet_name(workbook)
    if sheet_name is None:
        available = ", ".join(workbook.sheet_names) or "(none)"
        return None, f"Rate card sheet not found (available: {available})"

    preview = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        header=None,
        nrows=RA_HEADER_SCAN_ROWS,
    )
    header_row = find_ra_header_row(preview)
    if header_row is None:
        return None, "Rate card sheet is missing Origin/Destination header row"

    ra_df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
    ra_df.columns = [str(column).strip() for column in ra_df.columns]
    if ra_df.empty:
        return None, "Rate card sheet is empty"

    column_map = resolve_ra_column_map(ra_df)
    origin_column = column_map.get("Origin country")
    destination_column = column_map.get("Destination country")
    if origin_column is None or destination_column is None:
        headers = ", ".join(str(column) for column in ra_df.columns[:20])
        return None, f"Rate card sheet is missing Origin/Destination columns (headers: {headers})"

    lane_mask = (
        ra_df[origin_column].fillna("").astype(str).str.strip().ne("")
        & ra_df[destination_column].fillna("").astype(str).str.strip().ne("")
    )
    ra_df = ra_df.loc[lane_mask].copy()
    if ra_df.empty:
        return None, "Rate card sheet has no lane rows"

    return ra_df, ""


def resolve_ra_column_map(df: pd.DataFrame) -> dict[str, str]:
    column_names = {str(column).strip(): str(column) for column in df.columns}
    resolved: dict[str, str] = {}

    normalized_by_column = {
        column: normalize_header_name(column) for column in column_names
    }
    for canonical, aliases in RA_COLUMN_ALIASES.items():
        for column, normalized in normalized_by_column.items():
            if normalized in aliases:
                resolved[canonical] = column_names[column]
                break
    return resolved


def service_level_tokens(value: object) -> frozenset[str]:
    return frozenset(
        token
        for token in re.split(r"[,/;\s]+", cell_text(value).upper())
        if token
    )


def service_level_match_score(
    left_tokens: frozenset[str],
    right_tokens: frozenset[str],
) -> int:
    if not left_tokens and not right_tokens:
        return 1
    if not left_tokens or not right_tokens:
        return 0
    return len(left_tokens & right_tokens)


def ra_lane_base_key(
    *,
    tab_name: str,
    origin: str,
    destination: str,
    shipment_type: str,
    destination_postal_code: str,
) -> tuple[str, ...]:
    tab_name = cell_text(tab_name).upper()
    origin = cell_text(origin).upper()
    destination = cell_text(destination).upper()
    shipment_type = cell_text(shipment_type).upper()
    destination_postal_code = cell_text(destination_postal_code).upper()
    if is_sfs_tab(tab_name):
        return (tab_name, origin, destination, destination_postal_code, shipment_type)
    return (origin, destination, destination_postal_code, shipment_type)


def ra_lane_base_key_from_mapping(
    row: pd.Series,
    column_map: dict[str, str],
) -> tuple[str, ...] | None:
    origin_column = column_map.get("Origin country")
    destination_column = column_map.get("Destination country")
    if not origin_column or not destination_column:
        return None

    origin = cell_text(row.get(origin_column))
    destination = cell_text(row.get(destination_column))
    if not origin or not destination:
        return None

    tab_name = cell_text(row.get(column_map["Tab"])) if "Tab" in column_map else ""
    shipment_type = (
        cell_text(row.get(column_map["Shipment Type"]))
        if "Shipment Type" in column_map
        else ""
    )
    destination_postal_code = (
        cell_text(row.get(column_map["Destination Postal Code"]))
        if "Destination Postal Code" in column_map
        else ""
    )
    return ra_lane_base_key(
        tab_name=tab_name,
        origin=origin,
        destination=destination,
        shipment_type=shipment_type,
        destination_postal_code=destination_postal_code,
    )


def matrix_row_ra_base_key(row: pd.Series) -> tuple[str, ...]:
    return ra_lane_base_key(
        tab_name=cell_text(row.get("Tab")),
        origin=cell_text(row.get("Origin country")),
        destination=cell_text(row.get("Destination country")),
        shipment_type=cell_text(row.get("Shipment Type")),
        destination_postal_code=cell_text(row.get("Destination Postal Code")),
    )


def load_ra_lookup(
    file_path: Path,
) -> tuple[dict[tuple[str, ...], list[RaLaneEntry]] | None, frozenset[str] | str]:
    ra_df, read_error = read_ra_rate_card_dataframe(file_path)
    if ra_df is None:
        return None, read_error

    column_map = resolve_ra_column_map(ra_df)
    fill_fields = frozenset(field for field in RA_FILL_FIELDS if field in column_map)
    if not fill_fields:
        headers = ", ".join(str(column) for column in ra_df.columns)
        return None, f"Rate card sheet has no Category/Shipping Condition columns (headers: {headers})"

    lookup: dict[tuple[str, ...], list[RaLaneEntry]] = {}
    for _, row in ra_df.iterrows():
        base_key = ra_lane_base_key_from_mapping(row, column_map)
        if base_key is None:
            continue

        values: dict[str, str] = {}
        for field in fill_fields:
            value = cell_text(row.get(column_map[field]))
            if value:
                values[field] = value
        if not values:
            continue

        service_level = (
            cell_text(row.get(column_map["Service level"]))
            if "Service level" in column_map
            else ""
        )
        entry = RaLaneEntry(
            service_level_tokens=service_level_tokens(service_level),
            values=values,
        )
        lookup.setdefault(base_key, []).append(entry)

    if not lookup:
        return None, "Rate card sheet has no lane rows with Category/Shipping Condition values"
    return lookup, fill_fields


def match_ra_values_for_row(
    row: pd.Series,
    lookup: dict[tuple[str, ...], list[RaLaneEntry]],
) -> dict[str, str] | None:
    candidates = lookup.get(matrix_row_ra_base_key(row))
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0].values

    matrix_tokens = service_level_tokens(row.get("Service level"))
    best_entry = max(
        candidates,
        key=lambda entry: service_level_match_score(
            entry.service_level_tokens,
            matrix_tokens,
        ),
    )
    if service_level_match_score(best_entry.service_level_tokens, matrix_tokens) <= 0:
        return None
    return best_entry.values


def apply_ra_lookup(
    matrix_df: pd.DataFrame,
    lookup: dict[tuple[str, ...], list[RaLaneEntry]],
    *,
    fill_fields: frozenset[str],
) -> tuple[pd.DataFrame, set[tuple[int, str]]]:
    filled_cells: set[tuple[int, str]] = set()
    if matrix_df.empty or not lookup or not fill_fields:
        return matrix_df, filled_cells

    result = matrix_df.copy()
    for field in fill_fields:
        if field not in result.columns:
            result[field] = ""

    for position, (index, row) in enumerate(result.iterrows()):
        category = cell_text(row.get("Category"))
        if category == RETURN_CATEGORY:
            continue

        ra_values = match_ra_values_for_row(row, lookup)
        if ra_values is None:
            continue

        for field in fill_fields:
            value = ra_values.get(field, "")
            if not value:
                continue
            current = cell_text(result.at[index, field])
            if not current:
                result.at[index, field] = value
                filled_cells.add((position, field))

    return result, filled_cells


def normalize_business_segment_name(name: object) -> str:
    return re.sub(r"\s+", " ", cell_text(name).lower())


def business_segment_abbreviation(segment_name: object) -> str:
    normalized = normalize_business_segment_name(segment_name)
    if not normalized:
        return ""

    mapped = BUSINESS_SEGMENT_ABBREVIATIONS.get(normalized)
    if mapped:
        return mapped

    raw = cell_text(segment_name)
    if raw.isupper() and 1 < len(raw) <= 8:
        return raw

    words = re.findall(r"[a-z0-9]+", normalized)
    if words:
        return "".join(word[0] for word in words).upper()
    return raw


def load_billing_bs_lookup(file_path: Path | None = None) -> dict[str, str]:
    path = file_path or billing_bs_file_path()
    if path is None or not path.exists():
        return {}

    lookup: dict[str, str] = {}
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        if not line.strip():
            continue
        parts = line.split("\t")
        if len(parts) < 2:
            continue
        account_number = cell_text(parts[0])
        business_segment = cell_text(parts[1])
        if account_number and business_segment:
            lookup[account_number] = business_segment
    return lookup


def parse_carrier_account_numbers(
    value: object,
    billing_bs_lookup: dict[str, str] | None = None,
) -> list[str]:
    text = cell_text(value)
    if not text:
        return []

    tokens = CARRIER_ACCOUNT_TOKEN_PATTERN.findall(text)
    if billing_bs_lookup is None:
        return [
            token
            for part in CARRIER_ACCOUNT_SPLIT_PATTERN.split(text)
            for token in [re.split(r"[\s(]+", part.strip())[0].strip()]
            if token
        ]

    accounts: list[str] = []
    seen: set[str] = set()
    for token in tokens:
        if token not in billing_bs_lookup or token in seen:
            continue
        accounts.append(token)
        seen.add(token)
    return accounts


def resolve_business_segment(
    carrier_accounts: object,
    billing_bs_lookup: dict[str, str],
) -> str:
    abbreviations: list[str] = []
    seen: set[str] = set()

    for account_number in parse_carrier_account_numbers(
        carrier_accounts,
        billing_bs_lookup,
    ):
        segment_name = billing_bs_lookup.get(account_number)
        if not segment_name:
            continue
        abbreviation = business_segment_abbreviation(segment_name)
        if abbreviation and abbreviation not in seen:
            abbreviations.append(abbreviation)
            seen.add(abbreviation)

    return "/".join(abbreviations)


def apply_business_segment_column(
    matrix_df: pd.DataFrame,
    billing_bs_lookup: dict[str, str],
) -> pd.DataFrame:
    if matrix_df.empty or not billing_bs_lookup:
        return matrix_df

    result = matrix_df.copy()
    for index, row in result.iterrows():
        result.at[index, "Business Segment"] = resolve_business_segment(
            row.get("Carrier Account number"),
            billing_bs_lookup,
        )
    return result


def load_tab_index(file_path: Path) -> pd.DataFrame | None:
    workbook = pd.ExcelFile(file_path)
    for sheet_name in workbook.sheet_names:
        if sheet_name.strip().lower() == TAB_INDEX_SHEET.lower():
            return pd.read_excel(file_path, sheet_name=sheet_name)
    return None


def build_tab_index_lookup(tab_index_df: pd.DataFrame | None) -> dict[str, TabIndexInfo]:
    if tab_index_df is None or tab_index_df.empty:
        return {}

    lookup: dict[str, TabIndexInfo] = {}
    for _, row in tab_index_df.iterrows():
        tab_name = cell_text(row.get("Tab-Name"))
        if not tab_name:
            continue
        description = cell_text(row.get("Description"))
        status = cell_text(row.get("Status")).upper()
        lookup[tab_name] = TabIndexInfo(
            tab_name=tab_name,
            is_active=status == "ACTIVE",
            has_return_lanes=bool(RETURN_DESCRIPTION_PATTERN.search(description)),
            billing_accounts=normalize_billing_accounts(row.get("Billing Account(s)")),
            origins=cell_text(row.get("Origin(s)")),
        )
    return lookup


def apply_tab_index_metadata(
    matrix_df: pd.DataFrame,
    tab_index_lookup: dict[str, TabIndexInfo],
) -> pd.DataFrame:
    if matrix_df.empty or not tab_index_lookup:
        return matrix_df

    result = matrix_df.copy()
    if "Category" not in result.columns:
        result["Category"] = ""

    for index, row in result.iterrows():
        tab_name = cell_text(row.get("Tab"))
        tab_info = tab_index_lookup.get(tab_name)
        if tab_info is None:
            continue
        if tab_info.billing_accounts:
            result.at[index, "Carrier Account number"] = tab_info.billing_accounts

    return result


def append_return_lanes(
    matrix_df: pd.DataFrame,
    tab_index_lookup: dict[str, TabIndexInfo],
) -> pd.DataFrame:
    return_tabs = {
        tab_name
        for tab_name, tab_info in tab_index_lookup.items()
        if tab_info.has_return_lanes
    }
    if not return_tabs or matrix_df.empty:
        return matrix_df

    outbound = matrix_df[matrix_df["Category"].fillna("").astype(str).str.strip() == ""].copy()
    return_rows: list[dict[str, object]] = []

    for _, row in outbound.iterrows():
        if cell_text(row.get("Tab")) not in return_tabs:
            continue

        duplicate = row.to_dict()
        origin = cell_text(row.get("Origin country"))
        destination = cell_text(row.get("Destination country"))
        duplicate["Origin country"] = destination
        duplicate["Destination country"] = origin
        duplicate["Category"] = RETURN_CATEGORY
        return_rows.append(duplicate)

    if not return_rows:
        return matrix_df

    return pd.concat([matrix_df, pd.DataFrame(return_rows)], ignore_index=True)


def append_fr_time_definite_postal_lanes(matrix_df: pd.DataFrame) -> pd.DataFrame:
    """Duplicate FR Time Definite lanes with Destination Postal Code 98000."""
    if matrix_df.empty:
        return matrix_df

    destination_country = matrix_df["Destination country"].fillna("").astype(str).str.strip().str.upper()
    shipment_type = matrix_df["Shipment Type"].fillna("").astype(str)
    postal_code = matrix_df["Destination Postal Code"].fillna("").astype(str).str.strip()

    matching_mask = (
        destination_country.eq("FR")
        & shipment_type.str.contains(TIME_DEFINITE_PATTERN, na=False)
        & ~postal_code.eq(FR_TIME_DEFINITE_POSTAL_CODE)
    )
    matching = matrix_df[matching_mask]
    if matching.empty:
        return matrix_df

    duplicates: list[dict[str, object]] = []
    for _, row in matching.iterrows():
        duplicate = row.to_dict()
        duplicate["Destination Postal Code"] = FR_TIME_DEFINITE_POSTAL_CODE
        duplicates.append(duplicate)

    return pd.concat([matrix_df, pd.DataFrame(duplicates)], ignore_index=True)


def finalize_lane_numbers(matrix_df: pd.DataFrame) -> pd.DataFrame:
    if matrix_df.empty:
        return matrix_df
    result = matrix_df.copy()
    if "Category" in result.columns:
        result["Category"] = result["Category"].fillna("").astype(str).str.strip()
        result.loc[result["Category"].eq("nan"), "Category"] = ""
    result["Lane #"] = range(1, len(result) + 1)
    return result


def enrich_matrix_with_tab_index(
    matrix_df: pd.DataFrame,
    tab_index_df: pd.DataFrame | None,
    *,
    billing_bs_lookup: dict[str, str] | None = None,
) -> pd.DataFrame:
    tab_index_lookup = build_tab_index_lookup(tab_index_df)
    if billing_bs_lookup is None:
        billing_bs_lookup = load_billing_bs_lookup()
    result = apply_tab_index_metadata(matrix_df, tab_index_lookup)
    result = apply_business_segment_column(result, billing_bs_lookup)
    result = append_return_lanes(result, tab_index_lookup)
    result = apply_service_column(result, tab_index_lookup)
    return finalize_lane_numbers(result)


def collect_weight_brackets(
    rate_tabs: list[tuple[str, pd.DataFrame]],
    *,
    sfs_only: bool = False,
) -> tuple[list[str], dict[str, str]]:
    brackets: set[str] = set()
    bracket_rate_by: dict[str, str] = {}

    for tab_name, source_df in rate_tabs:
        if is_sfs_tab(tab_name) != sfs_only:
            continue
        for _, row in source_df.iterrows():
            bracket = weight_bracket_label(row.get("Chargeable Weight"))
            if not bracket:
                continue
            brackets.add(bracket)
            rate_by = rate_logic_to_rate_by(row.get("Rate Logic"))
            if rate_by:
                existing = bracket_rate_by.get(bracket)
                if existing and existing != rate_by:
                    scope = "SFS" if sfs_only else "Transport cost"
                    raise ValueError(
                        f"{scope} weight bracket {bracket} has conflicting rate logic: "
                        f"{existing} vs {rate_by}"
                    )
                bracket_rate_by[bracket] = rate_by

    ordered = sorted(brackets, key=bracket_sort_key)
    return ordered, bracket_rate_by


def collect_transport_cost_columns(
    rate_tabs: list[tuple[str, pd.DataFrame]],
) -> tuple[list[str], list[str], dict[str, str], dict[str, str]]:
    weight_brackets, bracket_rate_by = collect_weight_brackets(rate_tabs, sfs_only=False)
    sfs_brackets, sfs_bracket_rate_by = collect_weight_brackets(rate_tabs, sfs_only=True)
    standard_columns = [transport_cost_column_name(bracket) for bracket in weight_brackets]
    sfs_columns = [sfs_transport_cost_column_name(bracket) for bracket in sfs_brackets]
    return standard_columns, sfs_columns, bracket_rate_by, sfs_bracket_rate_by


def lane_key(
    origin: str,
    destination: str,
    shipment_type: str,
    stream_id: str,
    *,
    tab_name: str = "",
    destination_postal_code: str = "",
) -> tuple[str, ...]:
    if is_sfs_tab(tab_name):
        return (tab_name, origin, destination, destination_postal_code, shipment_type, stream_id)
    return (origin, destination, destination_postal_code, shipment_type, stream_id)


def row_service_level(source_row: pd.Series) -> str:
    return cell_text(source_row.get("Service Level")) or cell_text(source_row.get("Shipment Type"))


def build_shipment_and_cost_rows(
    rate_tabs: list[tuple[str, pd.DataFrame]],
    *,
    puk_mode: bool = False,
) -> pd.DataFrame:
    standard_columns, sfs_columns, _, _ = collect_transport_cost_columns(rate_tabs)
    transport_columns = [*standard_columns, *sfs_columns]
    lanes: dict[tuple[str, ...], dict[str, object]] = {}

    for tab_name, source_df in rate_tabs:
        dest_columns = destination_rate_columns(source_df, puk_mode=puk_mode)

        for _, source_row in source_df.iterrows():
            bracket = weight_bracket_label(source_row.get("Chargeable Weight"))
            if not bracket:
                continue

            origin = cell_text(source_row.get("Origin"))
            shipment_type = cell_text(source_row.get("Shipment Type"))
            service_level = row_service_level(source_row)
            default_destination_iso = cell_text(source_row.get("Destination ISO"))
            use_sfs_cost = is_sfs_tab(tab_name)
            cost_column = (
                sfs_transport_cost_column_name(bracket)
                if use_sfs_cost
                else transport_cost_column_name(bracket)
            )

            for dest_column in dest_columns:
                rate = rate_value(source_row.get(dest_column))
                if rate is None:
                    continue

                if puk_mode:
                    destination_country, destination_postal_code = parse_puk_destination_column(
                        dest_column,
                        default_destination_iso=default_destination_iso,
                    )
                else:
                    destination_country = dest_column
                    destination_postal_code = ""

                key = lane_key(
                    origin,
                    destination_country,
                    shipment_type,
                    service_level,
                    tab_name=tab_name,
                    destination_postal_code=destination_postal_code,
                )
                if key not in lanes:
                    lanes[key] = {
                        "Tab": tab_name,
                        "Origin country": origin,
                        "Shipment Type": shipment_type,
                        "Service": "",
                        "Service level": service_level,
                        "Destination country": destination_country,
                        "Destination Postal Code": destination_postal_code,
                        "Carrier Account number": "",
                        "Business Segment": "",
                        "Category": "",
                        "Shipping Condition": "",
                        "Valid from": format_display_date(source_row.get("Version Date")),
                        "Valid to": "",
                        CURRENCY_COLUMN: cell_text(source_row.get("Billing Currency")).upper(),
                        **{column: None for column in transport_columns},
                    }

                lane = lanes[key]
                if not lane["Valid from"]:
                    lane["Valid from"] = format_display_date(source_row.get("Version Date"))
                if not lane[CURRENCY_COLUMN]:
                    lane[CURRENCY_COLUMN] = cell_text(source_row.get("Billing Currency")).upper()

                existing_rate = lane.get(cost_column)
                if existing_rate is not None and existing_rate != rate:
                    raise ValueError(
                        f"Conflicting rates for lane {key} at {bracket}: {existing_rate} vs {rate}"
                    )
                lane[cost_column] = rate

    if not lanes:
        return pd.DataFrame(columns=[*SHIPMENT_COLUMNS, CURRENCY_COLUMN, *transport_columns])

    result = pd.DataFrame(list(lanes.values()))
    result["Lane #"] = range(1, len(result) + 1)
    columns = [*SHIPMENT_COLUMNS, CURRENCY_COLUMN, *transport_columns]
    return result[columns]


def _style_header_cell(
    cell,
    *,
    bold: bool = True,
    center: bool = False,
    fill: PatternFill = HEADER_FILL,
) -> None:
    cell.font = BOLD if bold else NORMAL
    cell.fill = fill
    cell.alignment = CENTER if center else LEFT
    cell.border = THIN_BORDER


def _style_cost_meta_cell(cell, *, fill: PatternFill = COST_META_FILL) -> None:
    cell.font = NORMAL
    cell.fill = fill
    cell.alignment = LEFT
    cell.border = THIN_BORDER


def _write_merged_header_cell(
    worksheet,
    row_index: int,
    start_col: int,
    end_col: int,
    value: str,
    *,
    fill: PatternFill,
    bold: bool = True,
    center: bool = True,
) -> None:
    if end_col > start_col:
        worksheet.merge_cells(
            start_row=row_index,
            start_column=start_col,
            end_row=row_index,
            end_column=end_col,
        )
    cell = worksheet.cell(row_index, start_col, value)
    _style_header_cell(cell, bold=bold, center=center, fill=fill)


def _write_merged_meta_cell(
    worksheet,
    row_index: int,
    start_col: int,
    end_col: int,
    value: str,
    *,
    fill: PatternFill = COST_META_FILL,
) -> None:
    if end_col > start_col:
        worksheet.merge_cells(
            start_row=row_index,
            start_column=start_col,
            end_row=row_index,
            end_column=end_col,
        )
    cell = worksheet.cell(row_index, start_col, value)
    _style_cost_meta_cell(cell, fill=fill)


def _column_width_for_header(header: str) -> float:
    return min(42.0, max(14.0, len(header) * 0.9 + 4))


def _write_transport_cost_group(
    worksheet,
    *,
    start_col: int,
    end_col: int,
    group_title: str,
    cost_columns: list[str],
    column_prefix: str,
    bracket_rate_by: dict[str, str],
    include_currency: bool = False,
) -> None:
    if not cost_columns:
        return

    _write_merged_header_cell(
        worksheet,
        COST_GROUP_ROW,
        start_col,
        end_col,
        group_title,
        fill=TRANSPORT_GROUP_FILL,
    )

    if include_currency:
        currency_name_cell = worksheet.cell(COST_NAME_ROW, start_col, CURRENCY_COLUMN)
        _style_header_cell(currency_name_cell, fill=TRANSPORT_COST_FILL, center=True)

    data_start_col = start_col + (1 if include_currency else 0)
    for offset, cost_column in enumerate(cost_columns, start=data_start_col):
        bracket_label = cost_column.removeprefix(column_prefix).removesuffix(")")
        cell = worksheet.cell(COST_NAME_ROW, offset, bracket_label)
        _style_header_cell(cell, fill=TRANSPORT_COST_FILL, center=True)

    _write_merged_meta_cell(
        worksheet,
        APPLY_IF_ROW,
        start_col,
        end_col,
        "Applies if invoiced by Carrier",
    )

    if include_currency:
        currency_rate_by_cell = worksheet.cell(RATE_BY_ROW, start_col)
        _style_cost_meta_cell(currency_rate_by_cell)

    for offset, cost_column in enumerate(cost_columns, start=data_start_col):
        bracket_label = cost_column.removeprefix(column_prefix).removesuffix(")")
        rate_by = bracket_rate_by.get(bracket_label, "")
        cell = worksheet.cell(
            RATE_BY_ROW,
            offset,
            f"Rate by: {rate_by}\nRegular rule" if rate_by else "",
        )
        _style_cost_meta_cell(cell)


def write_matrix_sheet(
    workbook: Workbook,
    matrix_df: pd.DataFrame,
    *,
    bracket_rate_by: dict[str, str],
    sfs_bracket_rate_by: dict[str, str] | None = None,
    sheet_name: str = "Rate card",
    ra_filled_cells: set[tuple[int, str]] | None = None,
) -> None:
    worksheet = workbook.active
    worksheet.title = sheet_name
    sfs_bracket_rate_by = sfs_bracket_rate_by or {}
    ra_filled_cells = ra_filled_cells or set()

    standard_columns = [
        column for column in matrix_df.columns if is_standard_weight_bracket_cost_column(column)
    ]
    sfs_columns = [column for column in matrix_df.columns if is_sfs_weight_bracket_cost_column(column)]
    shipment_count = len(SHIPMENT_COLUMNS)
    currency_col = shipment_count + 1
    standard_start_col = shipment_count + 2
    standard_end_col = shipment_count + 1 + len(standard_columns)
    sfs_start_col = standard_end_col + 1
    sfs_end_col = standard_end_col + len(sfs_columns)

    if standard_columns:
        _write_transport_cost_group(
            worksheet,
            start_col=currency_col,
            end_col=standard_end_col,
            group_title=f"Grouped cost: {TRANSPORT_COST_GROUP}",
            cost_columns=standard_columns,
            column_prefix=f"{TRANSPORT_COST_GROUP} (",
            bracket_rate_by=bracket_rate_by,
            include_currency=True,
        )

    if sfs_columns:
        _write_transport_cost_group(
            worksheet,
            start_col=sfs_start_col,
            end_col=sfs_end_col,
            group_title=f"Grouped cost: {SFS_TRANSPORT_COST_GROUP}",
            cost_columns=sfs_columns,
            column_prefix=f"{SFS_TRANSPORT_COST_GROUP} (",
            bracket_rate_by=sfs_bracket_rate_by,
            include_currency=False,
        )

    for col_index, header in enumerate(SHIPMENT_COLUMNS, start=1):
        cell = worksheet.cell(COLUMN_HEADER_ROW, col_index, header)
        _style_header_cell(cell, bold=header == "Lane #")

    if standard_columns:
        currency_header = worksheet.cell(COLUMN_HEADER_ROW, currency_col, CURRENCY_COLUMN)
        _style_header_cell(currency_header, center=True, fill=TRANSPORT_COST_FILL)

        for offset, cost_column in enumerate(standard_columns, start=standard_start_col):
            bracket_label = cost_column.removeprefix(f"{TRANSPORT_COST_GROUP} (").removesuffix(")")
            rate_by = bracket_rate_by.get(bracket_label, "Rate")
            header_cell = worksheet.cell(COLUMN_HEADER_ROW, offset, rate_by)
            _style_header_cell(header_cell, center=True, fill=TRANSPORT_COST_FILL)

    if sfs_columns:
        for offset, cost_column in enumerate(sfs_columns, start=sfs_start_col):
            bracket_label = cost_column.removeprefix(f"{SFS_TRANSPORT_COST_GROUP} (").removesuffix(")")
            rate_by = sfs_bracket_rate_by.get(bracket_label, "Rate")
            header_cell = worksheet.cell(COLUMN_HEADER_ROW, offset, rate_by)
            _style_header_cell(header_cell, center=True, fill=TRANSPORT_COST_FILL)

    all_cost_columns = [*standard_columns, *sfs_columns]
    for matrix_index, (_, row) in enumerate(matrix_df.iterrows()):
        excel_row = DATA_START_ROW + matrix_index

        for col_index, header in enumerate(SHIPMENT_COLUMNS, start=1):
            cell = worksheet.cell(excel_row, col_index, row.get(header))
            cell.alignment = LEFT
            cell.border = THIN_BORDER
            if (matrix_index, header) in ra_filled_cells:
                cell.fill = RA_FILL_HIGHLIGHT

        if standard_columns or sfs_columns:
            currency_cell = worksheet.cell(excel_row, currency_col, row.get(CURRENCY_COLUMN))
            currency_cell.alignment = CENTER
            currency_cell.border = THIN_BORDER

        for cost_column in all_cost_columns:
            offset = (
                currency_col + 1 + standard_columns.index(cost_column)
                if cost_column in standard_columns
                else sfs_start_col + sfs_columns.index(cost_column)
            )
            value = row.get(cost_column)
            cell = worksheet.cell(excel_row, offset)
            cell.border = THIN_BORDER
            if value is not None and value != "":
                cell.value = value
                cell.number_format = RATE_NUMBER_FORMAT
                cell.alignment = CENTER

    for col_index, header in enumerate(SHIPMENT_COLUMNS, start=1):
        worksheet.column_dimensions[get_column_letter(col_index)].width = _column_width_for_header(header)

    if standard_columns or sfs_columns:
        worksheet.column_dimensions[get_column_letter(currency_col)].width = 12.0

    for offset, cost_column in enumerate(standard_columns, start=standard_start_col):
        worksheet.column_dimensions[get_column_letter(offset)].width = _column_width_for_header(cost_column)

    for offset, cost_column in enumerate(sfs_columns, start=sfs_start_col):
        worksheet.column_dimensions[get_column_letter(offset)].width = _column_width_for_header(cost_column)

    worksheet.freeze_panes = worksheet.cell(DATA_START_ROW, 1)
    worksheet.sheet_view.showGridLines = False


def save_matrix(
    matrix_df: pd.DataFrame,
    source_file: Path,
    *,
    bracket_rate_by: dict[str, str],
    sfs_bracket_rate_by: dict[str, str] | None = None,
    output_path: Path | None = None,
    ra_filled_cells: set[tuple[int, str]] | None = None,
) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if output_path is None:
        output_path = OUTPUT_DIR / f"{source_file.stem.replace('_extracted', '')}_matrix.xlsx"

    workbook = Workbook()
    write_matrix_sheet(
        workbook,
        matrix_df,
        bracket_rate_by=bracket_rate_by,
        sfs_bracket_rate_by=sfs_bracket_rate_by,
        ra_filled_cells=ra_filled_cells,
    )
    workbook.save(output_path)
    return output_path


def run_build_matrix(
    *,
    source_file: Path | None = None,
    output_path: Path | None = None,
    auto: bool = False,
) -> Path:
    ensure_workspace_dirs()
    files = list_extracted_files()
    file_path = source_file
    if file_path is None:
        file_path = select_extracted_file(files, auto=auto)

    rate_tabs = load_rate_tabs(file_path)
    puk_mode = is_puk_rate_file(file_path)
    if not rate_tabs:
        raise RuntimeError(
            f"No active rate tabs found in {file_path.name}. "
            "Expected ACTIVE tabs from Tab Index with Origin, Chargeable Weight, "
            f"and destination ISO columns (excluding {', '.join(sorted(EXCLUDED_RATE_TABS))})."
        )

    print(f"\nBuilding matrix from {file_path.name}:")
    if puk_mode:
        print("  DHLPUK mode: postal code zones enabled (GB_NI, GB_HI, IE, ...)")
    for tab_name, tab_df in rate_tabs:
        print(f"  - {tab_name}: {len(tab_df)} source rows")

    _, _, bracket_rate_by, sfs_bracket_rate_by = collect_transport_cost_columns(rate_tabs)
    matrix_df = build_shipment_and_cost_rows(rate_tabs, puk_mode=puk_mode)
    tab_index_df = load_tab_index(file_path)
    billing_bs_lookup = load_billing_bs_lookup()
    matrix_df = enrich_matrix_with_tab_index(
        matrix_df,
        tab_index_df,
        billing_bs_lookup=billing_bs_lookup,
    )
    if is_eu_fr_postal_duplicate_file(file_path):
        before_count = len(matrix_df)
        matrix_df = append_fr_time_definite_postal_lanes(matrix_df)
        added = len(matrix_df) - before_count
        if added:
            matrix_df = finalize_lane_numbers(matrix_df)
            print(
                f"  FR Time Definite postal lanes added: {added} "
                f"(Destination Postal Code {FR_TIME_DEFINITE_POSTAL_CODE})"
            )

    ra_filled_cells: set[tuple[int, str]] = set()
    ra_file = select_ra_file(list_ra_files(), auto=auto)
    if ra_file is not None:
        ra_lookup, ra_fill_fields_or_reason = load_ra_lookup(ra_file)
        if ra_lookup is None:
            print(f"  RA lookup skipped for {ra_file.name}: {ra_fill_fields_or_reason}")
        else:
            ra_fill_fields = ra_fill_fields_or_reason
            before_filled = {
                field: matrix_df[field].fillna("").astype(str).str.strip().ne("").sum()
                for field in ra_fill_fields
                if field in matrix_df.columns
            }
            matrix_df, ra_filled_cells = apply_ra_lookup(
                matrix_df,
                ra_lookup,
                fill_fields=ra_fill_fields,
            )
            after_filled = {
                field: matrix_df[field].fillna("").astype(str).str.strip().ne("").sum()
                for field in ra_fill_fields
                if field in matrix_df.columns
            }
            filled_counts = {
                field: after_filled.get(field, 0) - before_filled.get(field, 0)
                for field in ra_fill_fields
            }
            ra_lane_count = sum(len(entries) for entries in ra_lookup.values())
            matched_lanes = sum(
                1
                for _, row in matrix_df.iterrows()
                if cell_text(row.get("Category")) != RETURN_CATEGORY
                and match_ra_values_for_row(row, ra_lookup) is not None
            )
            print(
                f"  RA lookup from {ra_file.name} (Rate card tab): "
                f"{ra_lane_count} RA lane(s), {matched_lanes} matched in result; filled "
                + ", ".join(f"{field}={filled_counts[field]}" for field in ra_fill_fields)
                + (f"; highlighted {len(ra_filled_cells)} cell(s)" if ra_filled_cells else "")
            )

    return_tab_names = [
        tab_name
        for tab_name, tab_info in build_tab_index_lookup(tab_index_df).items()
        if tab_info.has_return_lanes
    ]
    if return_tab_names:
        print(f"  Return lanes added for tabs: {', '.join(sorted(return_tab_names))}")
    if billing_bs_lookup:
        billing_bs_path = billing_bs_file_path()
        print(
            f"  Business Segment lookup: {len(billing_bs_lookup)} accounts "
            f"from {billing_bs_path.name if billing_bs_path else 'billing-bs'}"
        )
    if sfs_bracket_rate_by:
        print(
            f"  SFS transport cost brackets: "
            f"{', '.join(sfs_bracket_rate_by.keys())}"
        )

    saved_path = save_matrix(
        matrix_df,
        file_path,
        bracket_rate_by=bracket_rate_by,
        sfs_bracket_rate_by=sfs_bracket_rate_by,
        output_path=output_path,
        ra_filled_cells=ra_filled_cells,
    )
    if puk_mode:
        from build_postal_code_zones import run_build_postal_code_zones

        run_build_postal_code_zones(extracted_file=file_path)
    print(f"\nSaved matrix ({len(matrix_df)} lanes) to: {saved_path}")
    print(f"  Weight brackets: {', '.join(bracket_rate_by.keys())}")
    return saved_path


def main() -> int:
    try:
        run_build_matrix()
        return 0
    except KeyboardInterrupt:
        print("\nOperation cancelled.")
        return 130
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
